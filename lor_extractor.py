"""
=============================================================
  LOR — Extracción periódica del Top 1750 Maestros
=============================================================
  Optimizaciones incluidas:
    · Checkpoint guarda snapshot del leaderboard y PUUIDs resueltos
      → al reanudar retoma desde el índice exacto sin requests extra
    · Rate limiter de ventana deslizante (100 req / 2 min)
      → pausa proactiva antes de exceder la cuota, no después
    · Data Dragon pre-cargado al inicio para todos los sets conocidos
      → elimina "DD: sin maná" por sets no descargados a tiempo
    · XLSX se escribe cada 25 jugadores (CSV tras cada uno)
      → reduce I/O sin perder progreso

  Uso:
    python lor_extractor.py

  Requisitos:
    pip install requests openpyxl pandas
=============================================================
"""

import time
import json
import requests
import pandas as pd
from pathlib import Path
from collections import Counter, deque
from datetime import datetime, timezone
from openpyxl.styles import Font, PatternFill, Alignment


# ══════════════════════════════════════════════════════════════
# CONFIGURACIÓN  —  edita solo esta sección
# ══════════════════════════════════════════════════════════════

API_KEY       = "RGAPI-80e7fb25-30c1-4a57-8761-de0b337fa0bd"   # ← tu clave actual
REGION        = "americas"                      # americas | europe | sea
TOP_N         = 1750                            # jugadores a procesar
MATCHES_COUNT = 10                              # últimas N partidas por jugador
MAX_PENALTY   = 120                             # Retry-After > esto → detener

OUTPUT_CSV    = Path("lor_dataset.csv")
OUTPUT_XLSX   = Path("lor_dataset.xlsx")
CHECKPOINT    = Path("lor_checkpoint.json")

CANDIDATE_TAGS = ["NA1", "NA", "LAN", "LAS", "BR1", "BR",
                  "1", "001", "OCE1", "OCE", "1234", "0"]

# Sets de Data Dragon conocidos en LoR (se pre-cargan al inicio)
KNOWN_DD_SETS = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]


# ══════════════════════════════════════════════════════════════
# RATE LIMITER — ventana deslizante (100 req / 2 min)
# Llama a rl.wait() ANTES de cada request a la API de Riot.
# Data Dragon no cuenta contra este límite.
# ══════════════════════════════════════════════════════════════

class RateLimiter:
    # Ventana 1: 20 req/s  → usamos 18 como margen
    SEC_WINDOW = 1.0
    SEC_MAX    = 18
    # Ventana 2: 100 req/2min → usamos 95 como margen
    MIN_WINDOW = 120.0
    MIN_MAX    = 95

    def __init__(self):
        self._sec_ts: deque = deque()   # timestamps ventana 1s
        self._min_ts: deque = deque()   # timestamps ventana 120s

    def wait(self) -> None:
        while True:
            now = time.monotonic()

            # Limpiar timestamps fuera de cada ventana
            while self._sec_ts and now - self._sec_ts[0] >= self.SEC_WINDOW:
                self._sec_ts.popleft()
            while self._min_ts and now - self._min_ts[0] >= self.MIN_WINDOW:
                self._min_ts.popleft()

            sec_ok = len(self._sec_ts) < self.SEC_MAX
            min_ok = len(self._min_ts) < self.MIN_MAX

            if sec_ok and min_ok:
                break   # podemos hacer el request

            # Calcular cuánto esperar para la ventana más restrictiva
            if not min_ok:
                # La ventana de 2 min está llena → esperar hasta que
                # el request más antiguo salga de la ventana
                sleep_for = self.MIN_WINDOW - (now - self._min_ts[0]) + 0.05
                label = f"ventana 120s llena ({len(self._min_ts)}/{self.MIN_MAX})"
            else:
                # La ventana de 1s está llena → espera muy corta
                sleep_for = self.SEC_WINDOW - (now - self._sec_ts[0]) + 0.01
                label = f"ventana 1s llena ({len(self._sec_ts)}/{self.SEC_MAX})"

            if sleep_for > MAX_PENALTY:
                raise RateLimitPenalty(int(sleep_for))

            print(f"    ⏸  {label} — pausa {sleep_for:.2f}s")
            time.sleep(max(sleep_for, 0))

        # Registrar el request en ambas ventanas
        now = time.monotonic()
        self._sec_ts.append(now)
        self._min_ts.append(now)

    @property
    def used(self) -> tuple[int, int]:
        """Retorna (requests en ventana 1s, requests en ventana 120s)."""
        now = time.monotonic()
        sec = sum(1 for t in self._sec_ts if now - t < self.SEC_WINDOW)
        mn  = sum(1 for t in self._min_ts if now - t < self.MIN_WINDOW)
        return sec, mn


rl = RateLimiter()


# ══════════════════════════════════════════════════════════════
# CLIENTE DE LA API DE RIOT
# ══════════════════════════════════════════════════════════════

BASE_URL = f"https://{REGION}.api.riotgames.com"
HEADERS  = {"X-Riot-Token": API_KEY}


class RateLimitPenalty(Exception):
    def __init__(self, wait: int):
        self.wait = wait


def get(endpoint: str, params: dict = None, retries: int = 3):
    url = f"{BASE_URL}{endpoint}"
    for _ in range(retries):
        rl.wait()   # pausa proactiva antes del request
        r = requests.get(url, headers=HEADERS, params=params)
        if r.status_code == 200:
            return r.json()
        if r.status_code == 429:
            wait = int(r.headers.get("Retry-After", 10))
            if wait > MAX_PENALTY:
                raise RateLimitPenalty(wait)
            print(f"    ⏳ 429 recibido — esperando {wait}s...")
            time.sleep(wait)
            continue
        if r.status_code == 404:
            return None
        if r.status_code == 401:
            print("    ✗ 401 — API Key expirada. "
                  "Regénérala en developer.riotgames.com")
            raise SystemExit(1)
        print(f"    ✗ {r.status_code}: {r.text[:80]}")
        return None
    return None


# ══════════════════════════════════════════════════════════════
# DATA DRAGON — pre-carga completa al inicio
# ══════════════════════════════════════════════════════════════

_DD_CACHE: dict[int, dict] = {}   # {set_id: {card_code: card_data}}


def _load_set(set_id: int) -> bool:
    """
    Descarga el JSON de un set de Data Dragon.
    Retorna True si se cargó con datos, False si falló.
    No usa el rate limiter de Riot (Data Dragon es CDN pública).
    """
    if set_id in _DD_CACHE:
        return len(_DD_CACHE[set_id]) > 0

    url = (f"https://dd.b.pvp.net/latest/set{set_id}"
           f"/en_us/data/set{set_id}-en_us.json")
    try:
        r = requests.get(url, timeout=20)
        if r.status_code == 200:
            data = r.json()
            if data:
                _DD_CACHE[set_id] = {card["cardCode"]: card for card in data}
                return True
        _DD_CACHE[set_id] = {}
        return False
    except Exception as e:
        _DD_CACHE[set_id] = {}
        print(f"    ⚠  Set {set_id:02d}: {e}")
        return False


def preload_data_dragon() -> None:
    """Pre-carga todos los sets conocidos de Data Dragon al inicio."""
    print("[0/4] Pre-cargando Data Dragon...")
    loaded, failed = [], []
    for s in KNOWN_DD_SETS:
        ok = _load_set(s)
        if ok:
            loaded.append(s)
        else:
            failed.append(s)
    print(f"  ✓ Sets cargados : {loaded}")
    if failed:
        print(f"  ⚠  Sets no disponibles (las cartas de estos sets "
              f"tendrán mana_* = None): {failed}")
    total_cards = sum(len(v) for v in _DD_CACHE.values())
    print(f"  ✓ Total cartas en caché: {total_cards}\n")


def _get_cost(card_code: str) -> int | None:
    """Retorna el coste de maná real de una carta, o None si no disponible."""
    set_id = int(card_code[:2])
    # Si el set no está en caché (set desconocido), intentar descargarlo
    if set_id not in _DD_CACHE:
        _load_set(set_id)
    meta = _DD_CACHE.get(set_id, {}).get(card_code)
    return meta.get("cost") if meta else None


# ══════════════════════════════════════════════════════════════
# DECODIFICADOR DE DECK CODES  (spec abierta de Riot)
# https://github.com/RiotGames/LoRDeckCodes
# ══════════════════════════════════════════════════════════════

B32 = "ABCDEFGHIJKLMNOPQRSTUVWXYZ234567"

FACTION_ID = {
    0: "DE", 1: "FR", 2: "IO",  3: "NX",  4: "PZ",
    5: "SI", 6: "BW", 7: "SH",  9: "MT", 10: "BC", 12: "RU",
}
FACTION_NAME = {
    "DE": "Demacia",      "FR": "Freljord",        "IO": "Ionia",
    "NX": "Noxus",        "PZ": "Piltover & Zaun", "SI": "Shadow Isles",
    "BW": "Bilgewater",   "SH": "Shurima",         "MT": "Mount Targon",
    "BC": "Bandle City",  "RU": "Runeterra",
}


def _varint(data: list, idx: int) -> tuple[int, int]:
    r, s = 0, 0
    while idx < len(data):
        b = data[idx]; idx += 1
        r |= (b & 0x7F) << s
        if not (b & 0x80):
            break
        s += 7
    return r, idx


def decode(code: str) -> list[dict]:
    """Decodifica un deck_code → lista de {card_code, count, set, faction}."""
    bits = "".join(format(B32.index(c), "05b") for c in code.upper())
    raw  = [int(bits[i:i+8], 2) for i in range(0, (len(bits) // 8) * 8, 8)]
    cards, idx = [], 1

    for count in (3, 2):
        if idx >= len(raw):
            break
        ng, idx = _varint(raw, idx)
        for _ in range(ng):
            if idx >= len(raw): break
            nc, idx = _varint(raw, idx)
            if idx >= len(raw): break
            s,  idx = _varint(raw, idx)
            if idx >= len(raw): break
            f,  idx = _varint(raw, idx)
            fc = FACTION_ID.get(f, f"F{f:02d}")
            for _ in range(nc):
                if idx >= len(raw): break
                n, idx = _varint(raw, idx)
                cards.append({"card_code": f"{s:02d}{fc}{n:03d}",
                              "count": count, "set": s,
                              "faction": FACTION_NAME.get(fc, fc)})

    if idx < len(raw):
        ns, idx = _varint(raw, idx)
        for _ in range(ns):
            if idx >= len(raw): break
            s,  idx = _varint(raw, idx)
            if idx >= len(raw): break
            f,  idx = _varint(raw, idx)
            if idx >= len(raw): break
            n,  idx = _varint(raw, idx)
            fc = FACTION_ID.get(f, f"F{f:02d}")
            cards.append({"card_code": f"{s:02d}{fc}{n:03d}",
                          "count": 1, "set": s,
                          "faction": FACTION_NAME.get(fc, fc)})
    return cards


# ══════════════════════════════════════════════════════════════
# CONSTRUCCIÓN DE CAMPOS PARA EL DATASET
# ══════════════════════════════════════════════════════════════

def deck_to_row_fields(deck_code: str) -> dict:
    """
    Decodifica un deck_code y construye los campos del dataset.
    La curva de maná usa ÚNICAMENTE costes reales de Data Dragon.
    Si alguna carta no tiene coste disponible → mana_* y archetype_hint = None.
    """
    EMPTY = {"factions": None, "total_cards": None, "unique_cards": None,
             "card_list": None, "mana_low_pct": None,
             "mana_mid_pct": None, "mana_high_pct": None,
             "archetype_hint": None}

    try:
        cards = decode(deck_code)
    except Exception:
        return EMPTY

    if not cards:
        return EMPTY

    total = sum(c["count"] for c in cards)
    facs  = sorted({c["faction"] for c in cards})

    # Costes reales: si cualquier carta falta → toda la curva es None
    costs: list[int] = []
    dd_complete = True
    for c in cards:
        cost = _get_cost(c["card_code"])
        if cost is None:
            dd_complete = False
            break
        costs.extend([cost] * c["count"])

    if dd_complete and len(costs) == total:
        low  = sum(1 for x in costs if x <= 2)
        mid  = sum(1 for x in costs if 3 <= x <= 4)
        high = sum(1 for x in costs if x >= 5)
        lp   = round(low  / total * 100, 1)
        mp   = round(mid  / total * 100, 1)
        hp   = round(high / total * 100, 1)
        arch = "aggro" if lp > 50 else "control" if hp > 30 else "midrange"
    else:
        lp = mp = hp = arch = None

    return {
        "factions":       "|".join(facs),
        "total_cards":    total,
        "unique_cards":   len(cards),
        "card_list":      json.dumps(
            [{"c": c["card_code"], "n": c["count"]} for c in cards],
            separators=(",", ":"),
        ),
        "mana_low_pct":   lp,
        "mana_mid_pct":   mp,
        "mana_high_pct":  hp,
        "archetype_hint": arch,
    }


# ══════════════════════════════════════════════════════════════
# CHECKPOINT  —  guarda snapshot del leaderboard y PUUIDs
# ══════════════════════════════════════════════════════════════

def load_checkpoint() -> dict:
    if CHECKPOINT.exists():
        state = json.loads(CHECKPOINT.read_text())
        done  = state.get("last_index", 0)
        puuids = state.get("puuid_cache", {})
        print(f"  ✓ Checkpoint: índice {done}, "
              f"{len(puuids)} PUUIDs en caché, "
              f"{len(state.get('failed_names', []))} fallidos.")
        return state
    return {
        "last_index":   0,          # índice del último jugador procesado
        "players":      [],         # snapshot del leaderboard
        "puuid_cache":  {},         # {player_name: puuid}
        "failed_names": [],         # nombres sin PUUID encontrado
    }


def save_checkpoint(state: dict) -> None:
    CHECKPOINT.write_text(json.dumps(state, indent=2))


# ══════════════════════════════════════════════════════════════
# GUARDADO INCREMENTAL EN CSV + XLSX
# ══════════════════════════════════════════════════════════════

COLUMNS = [
    "extraction_date", "rank", "player_name", "lp",
    "puuid", "deck_code",
    "factions", "total_cards", "unique_cards",
    "mana_low_pct", "mana_mid_pct", "mana_high_pct",
    "archetype_hint", "card_list",
]


def load_existing() -> pd.DataFrame:
    if OUTPUT_CSV.exists():
        df = pd.read_csv(OUTPUT_CSV, dtype=str)
        print(f"  ✓ Dataset existente: {len(df)} filas.")
        return df
    return pd.DataFrame(columns=COLUMNS)


def save_csv(df: pd.DataFrame) -> None:
    df.to_csv(OUTPUT_CSV, index=False)


def save_xlsx(df: pd.DataFrame) -> None:
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="LOR_Masters")
        ws = writer.sheets["LOR_Masters"]

        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        header_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font      = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="left")

        ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════
# LÓGICA POR JUGADOR
# ══════════════════════════════════════════════════════════════

def resolve_puuid(name: str, puuid_cache: dict) -> str | None:
    """
    Retorna el PUUID del jugador.
    Si está en caché → sin requests.
    Si no → prueba CANDIDATE_TAGS y guarda el resultado en caché.
    """
    if name in puuid_cache:
        return puuid_cache[name]

    for tag in CANDIDATE_TAGS:
        data = get(f"/riot/account/v1/accounts/by-riot-id/{name}/{tag}")
        if data and "puuid" in data:
            puuid_cache[name] = data["puuid"]
            return data["puuid"]
    return None


def most_used_deck(puuid: str) -> str | None:
    """Retorna el deck_code más frecuente en las últimas N partidas."""
    ids = get(
        f"/lor/match/v1/matches/by-puuid/{puuid}/ids",
        params={"count": MATCHES_COUNT},
    )
    if not isinstance(ids, list) or not ids:
        return None

    counts: Counter = Counter()
    for mid in ids:
        match = get(f"/lor/match/v1/matches/{mid}")
        if not match:
            continue
        for p in match.get("info", {}).get("players", []):
            if p.get("puuid") == puuid:
                code = p.get("deck_code")
                if code:
                    counts[code] += 1
                break

    return counts.most_common(1)[0][0] if counts else None


# ══════════════════════════════════════════════════════════════
# EJECUCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════

def main() -> None:
    print("=" * 62)
    print("  LOR Extractor — Top 1750 Maestros")
    print("=" * 62)
    print(f"  Región     : {REGION}")
    print(f"  Jugadores  : {TOP_N}  |  Partidas: últimas {MATCHES_COUNT}")
    print(f"  Rate limit : {RateLimiter.SEC_MAX} req/s  |  "
    f"{RateLimiter.MIN_MAX} req/{RateLimiter.MIN_WINDOW:.0f}s (dual window)")
    print(f"  Outputs    : {OUTPUT_CSV}  +  {OUTPUT_XLSX}\n")

    # ── 0. Pre-cargar Data Dragon ─────────────────────────────────────────────
    preload_data_dragon()

    # ── Cargar estado ─────────────────────────────────────────────────────────
    state      = load_checkpoint()
    df         = load_existing()
    today      = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    puuid_cache: dict = state.get("puuid_cache", {})

    # ── 1. Leaderboard (solo si no hay snapshot en checkpoint) ────────────────
    if state.get("players"):
        players = state["players"]
        print(f"[1/4] Leaderboard desde checkpoint ({len(players)} jugadores).")
    else:
        print("[1/4] Obteniendo leaderboard Maestro...")
        lb = get("/lor/ranked/v1/leaderboards")
        if not lb or "players" not in lb:
            print("  ✗ No se pudo obtener el leaderboard. Verifica la API Key.")
            return
        players = lb["players"][:TOP_N]
        state["players"] = players
        save_checkpoint(state)
        print(f"  ✓ {len(players)} jugadores. Snapshot guardado en checkpoint.\n")

    # ── 2-4. Procesar jugadores desde el último índice ────────────────────────
    start_index = state.get("last_index", 0)
    pending     = players[start_index:]

    if not pending:
        print(f"  ✓ Todos los jugadores ya procesados en esta extracción.\n"
              f"    Borra {CHECKPOINT} para iniciar una nueva ronda.")
        return

    print(f"[2-4/4] Procesando jugadores {start_index + 1}–{len(players)}...")
    saved_count = 0

    for offset, player in enumerate(pending):
        i    = start_index + offset + 1   # posición real en el leaderboard (1-based)
        name = player.get("name", "")
        lp   = player.get("lp", 0)

        # Saltar si ya procesado HOY (en re-ejecuciones del mismo día)
        already_today = (
            not df.empty
            and ((df["player_name"] == name) & (df["extraction_date"] == today)).any()
        )
        if already_today:
            print(f"  [{i:>4}/{len(players)}] {name:<28} — ya extraído hoy.")
            state["last_index"] = i
            continue

        sec_used, min_used = rl.used
        print(f"  [{i:>4}/{len(players)}] {name:<28}  LP:{lp:<6} "
            f"[1s:{sec_used:>2}/{RateLimiter.SEC_MAX} | 2m:{min_used:>2}/{RateLimiter.MIN_MAX}]",
            end="", flush=True)

        try:
            # PUUID (desde caché si ya fue resuelto antes)
            puuid = resolve_puuid(name, puuid_cache)
            if not puuid:
                print("  ✗ sin PUUID")
                state["failed_names"].append(name)
                state["last_index"] = i
                state["puuid_cache"] = puuid_cache
                save_checkpoint(state)
                continue

            # Mazo más usado
            deck_code = most_used_deck(puuid)
            if not deck_code:
                print("  ✗ sin partidas")
                state["last_index"] = i
                state["puuid_cache"] = puuid_cache
                save_checkpoint(state)
                continue

            # Decodificar + enriquecer con Data Dragon
            fields = deck_to_row_fields(deck_code)

            # Construir fila
            row = {
                "extraction_date": today,
                "rank":            i,
                "player_name":     name,
                "lp":              lp,
                "puuid":           puuid,
                "deck_code":       deck_code,
                **fields,
            }
            new_row = pd.DataFrame([row], columns=COLUMNS)
            df      = pd.concat([df, new_row], ignore_index=True)
            saved_count += 1

            arch  = fields.get("archetype_hint") or "—"
            facs  = (fields.get("factions") or "—")[:28]
            mana  = "✓" if fields.get("mana_low_pct") is not None else "✗ sin maná"
            print(f"  {arch:<9} {facs:<28} DD:{mana}")

            # Actualizar checkpoint con nuevo índice y PUUIDs
            state["last_index"]  = i
            state["puuid_cache"] = puuid_cache
            save_csv(df)   # CSV tras cada jugador

            # XLSX y checkpoint cada 25 jugadores
            if saved_count % 25 == 0:
                save_xlsx(df)
                save_checkpoint(state)
                print(f"\n  💾 Checkpoint — posición {i}/{len(players)}, "
                      f"{len(df)} filas totales.\n")

        except RateLimitPenalty as e:
            state["last_index"]  = i - 1   # retrocede uno para reprocesar este
            state["puuid_cache"] = puuid_cache
            save_checkpoint(state)
            save_csv(df)
            save_xlsx(df)
            print(f"\n  🛑 Penalización larga ({e.wait}s ≈ {e.wait // 60} min).")
            print(f"     Datos guardados: {len(df)} filas.")
            print(f"     Espera {e.wait // 60} min y vuelve a ejecutar.")
            print(f"     Retomará desde el jugador {i} (posición guardada).")
            return

    # ── Guardado final ────────────────────────────────────────────────────────
    state["last_index"]  = len(players)
    state["puuid_cache"] = puuid_cache
    save_checkpoint(state)
    save_csv(df)
    save_xlsx(df)

    print(f"\n{'=' * 62}")
    print(f"  ✅ Extracción completada")
    print(f"     Nuevas filas esta ejecución : {saved_count}")
    print(f"     Total filas en el dataset   : {len(df)}")
    print(f"     Archivos: {OUTPUT_CSV}  |  {OUTPUT_XLSX}")
    print(f"\n  Para iniciar una nueva ronda borra: {CHECKPOINT}")
    print(f"{'=' * 62}\n")


if __name__ == "__main__":
    main()
